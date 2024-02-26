from selenium import webdriver
from bs4 import BeautifulSoup
import time
from datetime import datetime
import openpyxl
from openpyxl.styles import Font
def loop(month,year):
  # current_month = datetime.now().strftime('%B')
  # current_year=datetime.now().strftime('%Y')
  current_month=month
  current_year=year

  print(current_year)
  print(current_month)
  vulnerability_sheetname="vulnerabilities"
  opensource_sheetname='open-source'



  array=['CVE ID','Title','Description','Technology Area','Vulnerability Type','Access Vector','Security Rating','CVSS Rating','CVSS Score','CVSS String','Date Reported','Customer Notified Date','Affected Chipsets*','Patch**']
  app=openpyxl.load_workbook('incident.xlsx')

  vul_sheet=f"{current_month}-{current_year}-vulnerabilities"
  opensource_sheet=f"{current_month}-{current_year}-opensource"
  ws=app["Qualcomm"]
  

  driver = webdriver.Chrome()
  driver.get(f"https://docs.qualcomm.com/product/publicresources/securitybulletin/{current_month.lower()}-{current_year}-bulletin.html")

  time.sleep(20)
  page_source = driver.page_source
  soup = BeautifulSoup(page_source, "html.parser") 






  def all():      
      data_element = soup.find_all("section")
      for table in data_element:
          table_value=table.find_all('table')
          for tbody in table_value:
            tr=tbody.find_all('tbody')
            for findtr in tr:
              Value=findtr.find_all('tr')
              maxRow=ws.max_row
              for findtd in Value:
                final_data=findtd.find_all('td')
                         
                if(final_data!=None):
                   if(len(final_data)==2):
                        cool=final_data[0].get_text()
                        something=cool.strip()

                   
                        if(something=='CVE ID'):
                           ws.cell(maxRow+1,1).value=final_data[1].get_text()  
                           ws.cell(maxRow+1,15).value="Proprietary Software Issues"
                           ws.cell(maxRow+1,16).value=current_year
                           ws.cell(maxRow+1,17).value=current_month
                        if(something=='Title'):
                           ws.cell(maxRow+1,2).value=final_data[1].get_text()
                        if(something=='Description'):
                           ws.cell(maxRow+1,3).value=final_data[1].get_text()
                        if(something=='Technology Area'):
                           ws.cell(maxRow+1,4).value=final_data[1].get_text()
                        if(something=='Vulnerability Type'):
                           ws.cell(maxRow+1,5).value=final_data[1].get_text()
                        if(something=='Access Vector'):
                           ws.cell(maxRow+1,6).value=final_data[1].get_text()
                        if(something=='Security Rating'):
                           ws.cell(maxRow+1,7).value=final_data[1].get_text()
                        if(something=='CVSS Rating'):
                           ws.cell(maxRow+1,8).value=final_data[1].get_text()
                        if(something=='CVSS Score'):
                           ws.cell(maxRow+1,9).value=final_data[1].get_text()
                        if(something=='CVSS String'):
                           ws.cell(maxRow+1,10).value=final_data[1].get_text() 
                        if(something=='Date Reported'):
                           ws.cell(maxRow+1,11).value=final_data[1].get_text()
                        if(something=='Customer Notified Date'):
                           ws.cell(maxRow+1,12).value=final_data[1].get_text()
                        if(something=='Affected Chipsets*' or something=='Affected Chipsets' ):
                           ws.cell(maxRow+1,13).value=final_data[1].get_text()
                        if(something=="Patch**"):
                           ws.cell(maxRow+1,14).value=final_data[1].get_text()
                           ws.cell(maxRow+1,15).value="Open Source Software Issues"
                           
                           
               
                    
               
              

      # data=data_element.find_all('table')[2:]
      # for i in range(len(data)):
      #     value=data[i].find_all('tbody')
      #     for j in value:
      #       tr= j.find_all('tr')
      #       for k in tr:
      #           final= k.find_all('td')
      #           if(final[0].get_text()=="CVE ID"):
      #             ws.cell(i+2,1).value=final[1].get_text()
      #           if(final[0].get_text()=="Title"):
      #             ws.cell(i+2,2).value=final[1].get_text()
      #           if(final[0].get_text()=="Description"):
      #             ws.cell(i+2,3).value=final[1].get_text()
      #           if(final[0].get_text()=="Technology Area"):
      #             ws.cell(i+2,4).value=final[1].get_text()
      #           if(final[0].get_text()=="Vulnerability Type"):
      #             ws.cell(i+2,5).value=final[1].get_text()
      #           if(final[0].get_text()=="Access Vector"):
      #             ws.cell(i+2,6).value=final[1].get_text()
      #           if(final[0].get_text()=="Security Rating"):
      #             ws.cell(i+2,7).value=final[1].get_text()
      #           if(final[0].get_text()=="CVSS Rating"):
      #             ws.cell(i+2,8).value=final[1].get_text()
      #           if(final[0].get_text()=="CVSS Score"):
      #             ws.cell(i+2,9).value=final[1].get_text()
      #           if(final[0].get_text()=="CVSS String"):
      #             ws.cell(i+2,10).value=final[1].get_text()
      #           if(final[0].get_text()=="Date Reported"):
      #             ws.cell(i+2,11).value=final[1].get_text()
      #           if(final[0].get_text()=="Customer Notified Date"):
      #             ws.cell(i+2,12).value=final[1].get_text()
      #           if(final[0].get_text()=="Affected Chipsets*"):
      #             ws.cell(i+2,13).value=final[1].get_text()

  


  all()

  app.save('incident.xlsx')





year=['2019','2020','2021','2022','2023']

month=['January', 'February', 'March', 'April','May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']

#'January', 'February', 'March', 'April',


loop("February",'2024')

# driver = webdriver.Chrome()

# driver.get("https://docs.qualcomm.com/product/publicresources/securitybulletin/january-2024-bulletin.html")

# https://docs.qualcomm.com/product/publicresources/securitybulletin/february-2024-bulletin.html

# page_source = driver.page_source
# soup = BeautifulSoup(page_source, "html.parser")

# time.sleep(10)
# data_element = soup.find("section" ,class_="vulnerabilities")

# data=data_element.find_all('table')[2:]

# a=data[0].find_all('tr')
# print(len(a))








