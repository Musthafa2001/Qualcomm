from selenium import webdriver
from bs4 import BeautifulSoup
import time
from datetime import datetime
import openpyxl
from openpyxl.styles import Font
def loop(month,year):
  # current_month = datetime.now().strftime('%B')
  # current_year=datetime.now().strftime('%Y')
  current_month=month
  current_year=year

  print(current_year)
  print(current_month)
  vulnerability_sheetname="vulnerabilities"
  opensource_sheetname='open-source'



  array=['CVE ID','Title','Description','Technology Area','Vulnerability Type','Access Vector','Security Rating','CVSS Rating','CVSS Score','CVSS String','Date Reported','Customer Notified Date','Affected Chipsets*']
  open_source_list=['CVE ID','Title','Description','Technology Area','Vulnerability Type','Access Vector','Security Rating','CVSS Rating','CVSS Score','CVSS String','Date Reported','Customer Notified Date','Affected Chipsets*','Patch**']
  app=openpyxl.load_workbook('incident.xlsx')

  vul_sheet=f"{current_month}-{current_year}-vulnerabilities"
  opensource_sheet=f"{current_month}-{current_year}-opensource"




  driver = webdriver.Chrome()
  driver.get(f"https://docs.qualcomm.com/product/publicresources/securitybulletin/{current_month.lower()}-{current_year}-bulletin.html")

  time.sleep(10)
  page_source = driver.page_source
  soup = BeautifulSoup(page_source, "html.parser") 


  font = Font(bold=True)

  def qualcomm(): 
      ws=""
      if vul_sheet in app.sheetnames:
        ws=app[vul_sheet]
      else:
        ws=app.create_sheet(f"{current_month}-{current_year}-vulnerabilities")
        for i in range(len(array)):
          
          ws.cell(1,i+1).value=array[i]
          ws.cell(1,i+1).font=font
          
          print(ws.cell(1,i+1).value)
      data_element = soup.find("section" ,class_="vulnerabilities")
      data=data_element.find_all('table')[2:]
      for i in range(len(data)):
          value=data[i].find_all('tbody')
          for j in value:
            tr= j.find_all('tr')
            for k in tr:
                final= k.find_all('td')
                if(final[0].get_text()=="CVE ID"):
                  ws.cell(i+2,1).value=final[1].get_text()
                if(final[0].get_text()=="Title"):
                  ws.cell(i+2,2).value=final[1].get_text()
                if(final[0].get_text()=="Description"):
                  ws.cell(i+2,3).value=final[1].get_text()
                if(final[0].get_text()=="Technology Area"):
                  ws.cell(i+2,4).value=final[1].get_text()
                if(final[0].get_text()=="Vulnerability Type"):
                  ws.cell(i+2,5).value=final[1].get_text()
                if(final[0].get_text()=="Access Vector"):
                  ws.cell(i+2,6).value=final[1].get_text()
                if(final[0].get_text()=="Security Rating"):
                  ws.cell(i+2,7).value=final[1].get_text()
                if(final[0].get_text()=="CVSS Rating"):
                  ws.cell(i+2,8).value=final[1].get_text()
                if(final[0].get_text()=="CVSS Score"):
                  ws.cell(i+2,9).value=final[1].get_text()
                if(final[0].get_text()=="CVSS String"):
                  ws.cell(i+2,10).value=final[1].get_text()
                if(final[0].get_text()=="Date Reported"):
                  ws.cell(i+2,11).value=final[1].get_text()
                if(final[0].get_text()=="Customer Notified Date"):
                  ws.cell(i+2,12).value=final[1].get_text()
                if(final[0].get_text()=="Affected Chipsets*"):
                  ws.cell(i+2,13).value=final[1].get_text()

  def opensource():
    ws=""
    if opensource_sheet in app.sheetnames:
        ws=app[opensource_sheet]
    else:
        ws=app.create_sheet(f"{current_month}-{current_year}-opensource")
        for i in range(len(open_source_list)):
          a=ws.cell(1,i+1).value
          ws.cell(1,i+1).value=open_source_list[i]
          ws.cell(1,i+1).font=font
          print(ws.cell(1,i+1).value)
    data_element = soup.find("section" ,class_="open-source")
    data=data_element.find_all('table')[2:]
    for i in range(len(data)):
        value=data[i].find_all('tbody')
        for j in value:
            tr= j.find_all('tr')
            for k in tr:
              final= k.find_all('td')
              if(final[0].get_text()=="CVE ID"):
                ws.cell(i+2,1).value=final[1].get_text()
              if(final[0].get_text()=="Title"):
                ws.cell(i+2,2).value=final[1].get_text()
              if(final[0].get_text()=="Description"):
                ws.cell(i+2,3).value=final[1].get_text()
              if(final[0].get_text()=="Technology Area"):
                ws.cell(i+2,4).value=final[1].get_text()
              if(final[0].get_text()=="Vulnerability Type"):
                ws.cell(i+2,5).value=final[1].get_text()
              if(final[0].get_text()=="Access Vector"):
                ws.cell(i+2,6).value=final[1].get_text()
              if(final[0].get_text()=="Security Rating"):
                ws.cell(i+2,7).value=final[1].get_text()
              if(final[0].get_text()=="CVSS Rating"):
                ws.cell(i+2,8).value=final[1].get_text()
              if(final[0].get_text()=="CVSS Score"):
                ws.cell(i+2,9).value=final[1].get_text()
              if(final[0].get_text()=="CVSS String"):
                ws.cell(i+2,10).value=final[1].get_text()
              if(final[0].get_text()=="Date Reported"):
                ws.cell(i+2,11).value=final[1].get_text()
              if(final[0].get_text()=="Customer Notified Date"):
                ws.cell(i+2,12).value=final[1].get_text()
              if(final[0].get_text()=="Affected Chipsets*"):
                ws.cell(i+2,13).value=final[1].get_text() 
              if(final[0].get_text()=="Patch**"):
                ws.cell(i+2,14).value=final[1].get_text() 
    
  # Proprietary Software Issues 
  def test():
    ws=""
    if vul_sheet in app.sheetnames:
        ws=app[vul_sheet]
    else:
        ws=app.create_sheet(f"{current_month}-{current_year}-vulnerabilities")
        for i in range(len(array)):
          
          ws.cell(1,i+1).value=array[i]
          ws.cell(1,i+1).font=font
          
          print(ws.cell(1,i+1).value)


    data_element = soup.find_all("section")
    for j in range(len(data_element)):
       if(data_element[j].find('h2')!=None):
         if(data_element[j].find('h2').get_text()=="Proprietary Software Issues"):
           data=data_element[j].find_all('table')[2:]
           for i in range(len(data)):
               value=data[i].find_all('tbody')
               for j in value:
                  tr= j.find_all('tr')
                  for k in tr:
                    final= k.find_all('td')
                    if(final[0].get_text()=="CVE ID"):
                      ws.cell(i+2,1).value=final[1].get_text()
                    if(final[0].get_text()=="Title"):
                      ws.cell(i+2,2).value=final[1].get_text()
                    if(final[0].get_text()=="Description"):
                      ws.cell(i+2,3).value=final[1].get_text()
                    if(final[0].get_text()=="Technology Area"):
                      ws.cell(i+2,4).value=final[1].get_text()
                    if(final[0].get_text()=="Vulnerability Type"):
                      ws.cell(i+2,5).value=final[1].get_text()
                    if(final[0].get_text()=="Access Vector"):
                      ws.cell(i+2,6).value=final[1].get_text()
                    if(final[0].get_text()=="Security Rating"):
                      ws.cell(i+2,7).value=final[1].get_text()
                    if(final[0].get_text()=="CVSS Rating"):
                      ws.cell(i+2,8).value=final[1].get_text()
                    if(final[0].get_text()=="CVSS Score"):
                      ws.cell(i+2,9).value=final[1].get_text()
                    if(final[0].get_text()=="CVSS String"):
                      ws.cell(i+2,10).value=final[1].get_text()
                    if(final[0].get_text()=="Date Reported"):
                      ws.cell(i+2,11).value=final[1].get_text()
                    if(final[0].get_text()=="Customer Notified Date"):
                      ws.cell(i+2,12).value=final[1].get_text()
                    if(final[0].get_text()=="Affected Chipsets*"):
                      ws.cell(i+2,13).value=final[1].get_text() 


  def test2():
    ws=""
    if opensource_sheet in app.sheetnames:
        ws=app[opensource_sheet]
    else:
        ws=app.create_sheet(f"{current_month}-{current_year}-opensource")
        for i in range(len(open_source_list)):
          a=ws.cell(1,i+1).value
          ws.cell(1,i+1).value=open_source_list[i]
          ws.cell(1,i+1).font=font
          print(ws.cell(1,i+1).value)



    data_element = soup.find_all("section")
    for j in range(len(data_element)):
       if(data_element[j].find('h2')!=None):
         if(data_element[j].find('h2').get_text()=="Open Source Software Issues"):
           data=data_element[j].find_all('table')[2:]
           for i in range(len(data)):
               value=data[i].find_all('tbody')
               for j in value:
                  tr= j.find_all('tr')
                  for k in tr:
                    final= k.find_all('td')
                    if(final[0].get_text()=="CVE ID"):
                      ws.cell(i+2,1).value=final[1].get_text()
                    if(final[0].get_text()=="Title"):
                      ws.cell(i+2,2).value=final[1].get_text()
                    if(final[0].get_text()=="Description"):
                      ws.cell(i+2,3).value=final[1].get_text()
                    if(final[0].get_text()=="Technology Area"):
                      ws.cell(i+2,4).value=final[1].get_text()
                    if(final[0].get_text()=="Vulnerability Type"):
                      ws.cell(i+2,5).value=final[1].get_text()
                    if(final[0].get_text()=="Access Vector"):
                      ws.cell(i+2,6).value=final[1].get_text()
                    if(final[0].get_text()=="Security Rating"):
                      ws.cell(i+2,7).value=final[1].get_text()
                    if(final[0].get_text()=="CVSS Rating"):
                      ws.cell(i+2,8).value=final[1].get_text()
                    if(final[0].get_text()=="CVSS Score"):
                      ws.cell(i+2,9).value=final[1].get_text()
                    if(final[0].get_text()=="CVSS String"):
                      ws.cell(i+2,10).value=final[1].get_text()
                    if(final[0].get_text()=="Date Reported"):
                      ws.cell(i+2,11).value=final[1].get_text()
                    if(final[0].get_text()=="Customer Notified Date"):
                      ws.cell(i+2,12).value=final[1].get_text()
                    if(final[0].get_text()=="Affected Chipsets*"):
                      ws.cell(i+2,13).value=final[1].get_text() 
                    if(final[0].get_text()=="Patch**"):
                      ws.cell(i+2,14).value=final[1].get_text()              
    
      
      
     
      
      

      # if(i.find('h2').get_text()=='Proprietary Software Issues'):
      #   print(i.find('table'))

  


          


  # qualcomm()

  # opensource()
                
  test()#2022
  test2()#2022

  






  app.save('incident.xlsx')



year=['2019','2020','2021','2022','2023']

month=['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']



for mon in month:
  loop(mon,'2021')
# driver = webdriver.Chrome()

# driver.get("https://docs.qualcomm.com/product/publicresources/securitybulletin/january-2024-bulletin.html")

# https://docs.qualcomm.com/product/publicresources/securitybulletin/february-2024-bulletin.html

# page_source = driver.page_source
# soup = BeautifulSoup(page_source, "html.parser")

# time.sleep(10)
# data_element = soup.find("section" ,class_="vulnerabilities")

# data=data_element.find_all('table')[2:]

# a=data[0].find_all('tr')
# print(len(a))








